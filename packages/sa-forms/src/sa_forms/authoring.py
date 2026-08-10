"""Build a form definition from an existing artifact.

The user uploads the Excel sheet (or CSV, or JSON) they fill in today, and this
produces a working :class:`FormDefinition` instead of them authoring one by
hand.

Two stages, split on purpose:

1. **Structural inspection** — deterministic. Read the sheet, find the header
   row, infer each column's type from its sample values, spot enum columns from
   low-cardinality repeats, guess required-ness from fill rate. This is
   arithmetic, and a model would only make it less reliable.
2. **Semantic enrichment** — the model's job. Turn ``req_dt`` into "Requested
   date", write the description and the *rationale* (the "why do you need
   this?" answer), and propose aliases people would actually say.

Then the facilitator asks about whatever remains genuinely ambiguous, and only
then is the form registered — as a ``DRAFT``, never auto-activated.
"""

from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass
from dataclasses import field as dataclass_field
from pathlib import Path
from typing import Any

from sa_platform.errors import ValidationError
from sa_platform.logging import get_logger

from .models import (
    FieldType,
    FormDefinition,
    FormField,
    FormSection,
    FormStatus,
    Importance,
)

logger = get_logger(__name__)

#: A column whose values repeat, drawn from a small set, is a picklist.
_ENUM_MAX_DISTINCT = 12
#: Tuned against small samples: three values across six rows (0.5) is a
#: picklist, twenty distinct names across twenty rows (1.0) is free text.
_ENUM_MAX_RATIO = 0.6
#: A column filled this often in the sample is probably mandatory.
_REQUIRED_FILL_RATIO = 0.9

_DATE_HINT = re.compile(r"\b(date|deadline|due|when|schedule|start|end|go.?live)\b", re.I)
_BOOL_HINT = re.compile(r"^(is|has|needs?|requires?|should|can|enable)_|\?$", re.I)
_PERSON_HINT = re.compile(
    r"\b(owner|approver|requester|assignee|contact|manager|lead|author)\b", re.I
)
_EMAIL_HINT = re.compile(r"\b(e-?mail|mail)\b", re.I)
_URL_HINT = re.compile(r"\b(url|link|href|website)\b", re.I)
_MONEY_HINT = re.compile(r"\b(cost|budget|price|amount|spend|value|fee)\b", re.I)
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}")
_NUMERIC = re.compile(r"^-?[\d,]+\.?\d*$")


@dataclass(slots=True)
class ColumnProfile:
    """What structural inspection learned about one column."""

    name: str
    samples: list[str] = dataclass_field(default_factory=list)
    filled: int = 0
    total: int = 0
    distinct: list[str] = dataclass_field(default_factory=list)
    inferred_type: FieldType = FieldType.STRING
    inferred_importance: Importance = Importance.OPTIONAL
    options: list[str] = dataclass_field(default_factory=list)
    #: Set when inspection could not settle something a human should confirm.
    open_questions: list[str] = dataclass_field(default_factory=list)

    @property
    def fill_ratio(self) -> float:
        return self.filled / self.total if self.total else 0.0


@dataclass(slots=True)
class InferenceReport:
    """The draft form plus what the facilitator should ask about."""

    definition: FormDefinition
    columns: list[ColumnProfile] = dataclass_field(default_factory=list)
    questions: list[str] = dataclass_field(default_factory=list)
    warnings: list[str] = dataclass_field(default_factory=list)
    source: str = ""

    def summary(self) -> dict[str, Any]:
        return {
            "form": self.definition.name,
            "version": self.definition.version,
            "fields": self.definition.field_count(),
            "sections": len(self.definition.sections),
            "questions": self.questions,
            "warnings": self.warnings,
            "source": self.source,
        }


# ---------------------------------------------------------------------------
# Stage 1 — structural inspection
# ---------------------------------------------------------------------------


def _slug(text: str) -> str:
    """Turn a column heading into a valid field id."""
    cleaned = re.sub(r"[^\w\s]", " ", str(text)).strip().lower()
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned:
        return "field"
    if not cleaned[0].isalpha():
        cleaned = f"f_{cleaned}"
    return cleaned[:60]


def _infer_type(profile: ColumnProfile) -> FieldType:
    """Infer a column's type from its name and its values.

    Values win where they are unambiguous; the name only breaks ties. A column
    called `start_date` holding "Q3" is text, not a date.
    """
    name = profile.name
    values = [v for v in profile.samples if v.strip()]

    if values:
        if all(_ISO_DATE.match(v) for v in values):
            return FieldType.DATE
        if all("@" in v and "." in v for v in values):
            return FieldType.EMAIL
        if all(v.lower().startswith(("http://", "https://")) for v in values):
            return FieldType.URL
        lowered = {v.strip().lower() for v in values}
        if lowered <= {"yes", "no", "y", "n", "true", "false", "0", "1"}:
            return FieldType.BOOLEAN
        if all(_NUMERIC.match(v) for v in values):
            if _MONEY_HINT.search(name):
                return FieldType.CURRENCY
            return FieldType.INTEGER if all("." not in v for v in values) else FieldType.NUMBER

    if profile.options:
        return FieldType.ENUM
    if _EMAIL_HINT.search(name):
        return FieldType.EMAIL
    if _URL_HINT.search(name):
        return FieldType.URL
    if _DATE_HINT.search(name):
        return FieldType.DATE
    if _BOOL_HINT.search(name):
        return FieldType.BOOLEAN
    if _PERSON_HINT.search(name):
        return FieldType.PERSON
    if _MONEY_HINT.search(name):
        return FieldType.CURRENCY

    # Long free text is a different question shape from a short label.
    if values and sum(len(v) for v in values) / len(values) > 120:
        return FieldType.TEXT
    return FieldType.STRING


def profile_columns(headers: list[str], rows: list[list[str]]) -> list[ColumnProfile]:
    """Inspect each column's sample values."""
    profiles: list[ColumnProfile] = []

    for index, header in enumerate(headers):
        values = [
            str(row[index]).strip() if index < len(row) and row[index] is not None else ""
            for row in rows
        ]
        filled = [v for v in values if v and v.lower() not in ("n/a", "na", "-", "none")]

        profile = ColumnProfile(
            name=str(header).strip(),
            samples=filled[:20],
            filled=len(filled),
            total=len(values),
        )

        distinct = sorted(set(filled))
        profile.distinct = distinct[:50]

        # Low-cardinality repeated values are a picklist, not free text.
        if (
            1 < len(distinct) <= _ENUM_MAX_DISTINCT
            and len(filled) >= 3
            # Repetition is the signal: distinct == filled means every row is
            # unique, which is a name or a description, not a picklist.
            and len(distinct) < len(filled)
            and len(distinct) / max(len(filled), 1) <= _ENUM_MAX_RATIO
        ):
            profile.options = distinct

        profile.inferred_type = _infer_type(profile)
        if profile.inferred_type is not FieldType.ENUM:
            profile.options = []

        if profile.total == 0:
            profile.inferred_importance = Importance.RECOMMENDED
            profile.open_questions.append(
                f"'{profile.name}' had no sample values — is it required?"
            )
        elif profile.fill_ratio >= _REQUIRED_FILL_RATIO:
            profile.inferred_importance = Importance.MANDATORY
        elif profile.fill_ratio >= 0.4:
            profile.inferred_importance = Importance.RECOMMENDED
        else:
            profile.inferred_importance = Importance.OPTIONAL

        if profile.inferred_type is FieldType.ENUM and len(distinct) > _ENUM_MAX_DISTINCT:
            profile.open_questions.append(
                f"'{profile.name}' looks like a picklist but has {len(distinct)} values — "
                "should it be free text?"
            )

        profiles.append(profile)

    return profiles


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------


def read_tabular(content: bytes | str, *, filename: str = "") -> tuple[list[str], list[list[str]]]:
    """Read headers and rows from an uploaded artifact.

    Supports .xlsx/.xlsm via openpyxl, and CSV/TSV otherwise. A key-value sheet
    (two columns, ``Label | Value``) is transposed into one row, because that
    is how single-instance forms are usually laid out.
    """
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        return _read_excel(content if isinstance(content, bytes) else content.encode("utf-8"))
    if suffix == ".json":
        return _read_json(content if isinstance(content, str) else content.decode("utf-8"))
    return _read_delimited(content if isinstance(content, str) else content.decode("utf-8-sig"))


def _read_excel(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise ValidationError(
            "reading .xlsx samples requires openpyxl (install sa-forms[xlsx])", cause=exc
        ) from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    grid = [
        ["" if cell is None else str(cell).strip() for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    workbook.close()

    grid = [row for row in grid if any(cell for cell in row)]
    if not grid:
        raise ValidationError("the uploaded spreadsheet is empty")

    header_index = _find_header_row(grid)
    headers = grid[header_index]
    rows = grid[header_index + 1 :]

    # Trim trailing all-blank columns produced by stray formatting.
    width = max((i + 1 for i, h in enumerate(headers) if h), default=0)
    headers = headers[:width]
    rows = [row[:width] for row in rows]

    if _looks_like_key_value(headers, rows):
        return _transpose_key_value(rows)
    return headers, rows


def _find_header_row(grid: list[list[str]]) -> int:
    """Pick the header row.

    Real templates carry a title and blank spacer rows above the headings, so
    row 0 is often wrong. The header is the earliest row with the most
    non-empty, short, non-numeric cells.
    """
    best_index, best_score = 0, -1.0
    for index, row in enumerate(grid[:10]):
        cells = [c for c in row if c]
        if len(cells) < 2:
            continue
        shortish = sum(1 for c in cells if len(c) <= 60)
        non_numeric = sum(1 for c in cells if not _NUMERIC.match(c))
        score = len(cells) + shortish * 0.5 + non_numeric * 0.5 - index * 0.4
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _looks_like_key_value(headers: list[str], rows: list[list[str]]) -> bool:
    """Detect a vertical ``Label | Value`` layout."""
    if len(headers) != 2 or len(rows) < 2:
        return False
    labels = {h.strip().lower() for h in headers}
    return bool(labels & {"field", "label", "question", "item", "attribute", "key", "name"})


def _transpose_key_value(rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    headers = [row[0].strip() for row in rows if row and row[0].strip()]
    values = [row[1].strip() if len(row) > 1 else "" for row in rows if row and row[0].strip()]
    return headers, [values]


def _read_delimited(text: str) -> tuple[list[str], list[list[str]]]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    grid = [row for row in reader if any(cell.strip() for cell in row)]
    if not grid:
        raise ValidationError("the uploaded file contains no rows")

    header_index = _find_header_row(grid)
    return [c.strip() for c in grid[header_index]], grid[header_index + 1 :]


def _read_json(text: str) -> tuple[list[str], list[list[str]]]:
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValidationError(f"the uploaded JSON could not be parsed: {exc}", cause=exc) from exc

    records = payload if isinstance(payload, list) else [payload]
    records = [r for r in records if isinstance(r, dict)]
    if not records:
        raise ValidationError("the uploaded JSON contains no objects")

    headers: list[str] = []
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(key)

    rows = [
        ["" if record.get(h) is None else str(record.get(h)) for h in headers] for record in records
    ]
    return headers, rows


# ---------------------------------------------------------------------------
# Stage 2 — semantic enrichment
# ---------------------------------------------------------------------------

_ENRICH_SYSTEM = """\
You turn raw spreadsheet columns into a well-described form definition.

For every column produce:
- `label`: a clear human-readable name (expand abbreviations: `req_dt` -> \
"Requested date").
- `description`: one sentence saying what the field means.
- `rationale`: one sentence on WHY the form needs it and what it unblocks. This \
is shown verbatim when a user asks "why are you asking me this?", so make it a \
real reason, not a restatement of the label.
- `aliases`: other words people would use for it in conversation.
- `section`: a short topic name grouping related columns.

Also propose a `form_title` and a one-line `form_description`.

Group columns into 2-6 coherent sections. Preserve the given field_ids exactly. \
Do not invent, drop, merge, or rename fields.\
"""

_ENRICH_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "form_title": {"type": "string"},
        "form_description": {"type": "string"},
        "sections": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "id": {"type": "string"},
                    "title": {"type": "string"},
                    "description": {"type": "string"},
                    "field_ids": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["id", "title", "field_ids"],
                "additionalProperties": False,
            },
        },
        "fields": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "field_id": {"type": "string"},
                    "label": {"type": "string"},
                    "description": {"type": "string"},
                    "rationale": {"type": "string"},
                    "aliases": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["field_id", "label", "description", "rationale"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["form_title", "sections", "fields"],
    "additionalProperties": False,
}


class FormAuthor:
    """Infers form definitions from sample artifacts."""

    def __init__(self, llm_provider: Any | None = None) -> None:
        self._llm = llm_provider

    def _provider(self) -> Any:
        if self._llm is None:
            from sa_connectors.llm import build_provider

            self._llm = build_provider()
        return self._llm

    async def infer(
        self,
        content: bytes | str,
        *,
        filename: str = "sample.xlsx",
        form_name: str | None = None,
        enrich: bool = True,
    ) -> InferenceReport:
        """Produce a draft form definition from an uploaded sample."""
        headers, rows = read_tabular(content, filename=filename)
        if not headers:
            raise ValidationError("no column headings could be identified in the sample")

        profiles = profile_columns(headers, rows)
        profiles = self._deduplicate(profiles)

        name = form_name or _slug(Path(filename).stem) or "imported_form"
        report = InferenceReport(
            definition=self._build_baseline(name, filename, profiles),
            columns=profiles,
            source=filename,
        )

        if enrich:
            try:
                report.definition = await self._enrich(report.definition, profiles, rows)
            except Exception as exc:  # noqa: BLE001 - the structural draft still stands
                logger.error(
                    "form enrichment failed; keeping the structural draft",
                    extra={"form": name, "error": str(exc)},
                )
                report.warnings.append(
                    "Could not generate labels and rationales automatically; "
                    "the field names come straight from the sample."
                )

        report.questions = self._facilitator_questions(report.definition, profiles, rows)
        if not rows:
            report.warnings.append(
                "The sample had headings but no data rows, so field types and "
                "required-ness are guesses from the column names alone."
            )

        logger.info(
            "inferred form from sample",
            extra={"form": name, "fields": len(profiles), "source": filename},
        )
        return report

    @staticmethod
    def _deduplicate(profiles: list[ColumnProfile]) -> list[ColumnProfile]:
        """Ensure slugged ids stay unique — real sheets repeat headings."""
        seen: dict[str, int] = {}
        for profile in profiles:
            base = _slug(profile.name)
            if base in seen:
                seen[base] += 1
                profile.name = f"{profile.name} {seen[base]}"
            else:
                seen[base] = 1
        return [p for p in profiles if p.name.strip()]

    @staticmethod
    def _build_baseline(name: str, filename: str, profiles: list[ColumnProfile]) -> FormDefinition:
        """The structural draft, before any model involvement."""
        fields = [
            FormField(
                id=_slug(p.name),
                label=p.name,
                type=p.inferred_type,
                importance=p.inferred_importance,
                options=p.options,
                examples=p.samples[:3],
                description="",
                rationale="",
            )
            for p in profiles
        ]
        return FormDefinition(
            name=name,
            version="0.1.0",
            title=name.replace("_", " ").title(),
            description=f"Imported from {filename}.",
            status=FormStatus.DRAFT,
            derived_from=filename,
            sections=[
                FormSection(
                    id="details",
                    title="Details",
                    description="Fields imported from the sample.",
                    fields=fields,
                    order=0,
                )
            ],
        )

    async def _enrich(
        self,
        draft: FormDefinition,
        profiles: list[ColumnProfile],
        rows: list[list[str]],
    ) -> FormDefinition:
        """Add labels, descriptions, rationales, aliases, and sections."""
        from sa_connectors.llm.base import Message

        lines = [
            f"Sample file: {draft.derived_from}",
            f"{len(rows)} data row(s).",
            "",
            "Columns:",
        ]
        for profile in profiles:
            samples = "; ".join(profile.samples[:3]) or "(no sample values)"
            detail = (
                f'- field_id `{_slug(profile.name)}` | heading "{profile.name}" | '
                f"type {profile.inferred_type.value} | "
                f"filled {profile.fill_ratio:.0%} | examples: {samples}"
            )
            if profile.options:
                detail += f" | picklist: {', '.join(profile.options)}"
            lines.append(detail)

        enriched = await self._provider().complete_structured(
            [Message.user("\n".join(lines))], _ENRICH_SCHEMA, system=_ENRICH_SYSTEM
        )

        by_id = {_slug(p.name): p for p in profiles}
        metadata = {
            item["field_id"]: item
            for item in enriched.get("fields", [])
            if item.get("field_id") in by_id
        }

        # Build sections from the model's grouping, but drive membership from
        # our own id list so no field can be dropped or invented.
        assigned: set[str] = set()
        sections: list[FormSection] = []

        for order, raw_section in enumerate(enriched.get("sections", [])):
            member_ids = [
                fid
                for fid in raw_section.get("field_ids", [])
                if fid in by_id and fid not in assigned
            ]
            if not member_ids:
                continue
            assigned.update(member_ids)
            sections.append(
                FormSection(
                    id=_slug(raw_section.get("id", f"section_{order}")),
                    title=raw_section.get("title", f"Section {order + 1}"),
                    description=raw_section.get("description", ""),
                    order=order,
                    fields=[self._build_field(by_id[fid], metadata.get(fid)) for fid in member_ids],
                )
            )

        leftover = [fid for fid in by_id if fid not in assigned]
        if leftover:
            sections.append(
                FormSection(
                    id="other",
                    title="Other details",
                    description="Fields not grouped elsewhere.",
                    order=len(sections),
                    fields=[self._build_field(by_id[fid], metadata.get(fid)) for fid in leftover],
                )
            )

        if not sections:
            return draft

        return draft.model_copy(
            update={
                "title": enriched.get("form_title") or draft.title,
                "description": enriched.get("form_description") or draft.description,
                "sections": sections,
            }
        )

    @staticmethod
    def _build_field(profile: ColumnProfile, meta: dict[str, Any] | None) -> FormField:
        meta = meta or {}
        return FormField(
            id=_slug(profile.name),
            label=meta.get("label") or profile.name,
            type=profile.inferred_type,
            importance=profile.inferred_importance,
            description=meta.get("description", ""),
            rationale=meta.get("rationale", ""),
            aliases=[a for a in meta.get("aliases", []) if a][:6],
            options=profile.options,
            examples=profile.samples[:3],
            # A wrong date or owner is expensive to unwind, so require these to
            # be stated rather than inferred.
            require_explicit=profile.inferred_type
            in (FieldType.DATE, FieldType.DATETIME, FieldType.CURRENCY),
        )

    @staticmethod
    def _facilitator_questions(
        definition: FormDefinition,
        profiles: list[ColumnProfile],
        rows: list[list[str]],
    ) -> list[str]:
        """What the facilitator should confirm before the form goes live.

        Only genuine ambiguities. Asking about everything would recreate the
        manual authoring this feature exists to remove.
        """
        questions: list[str] = []
        by_id = {_slug(p.name): p for p in profiles}

        mandatory = [f for f in definition.fields() if f.is_mandatory]
        if mandatory:
            labels = ", ".join(f.label for f in mandatory[:8])
            more = f" (+{len(mandatory) - 8} more)" if len(mandatory) > 8 else ""
            questions.append(
                f"I marked these as required based on how consistently they were filled "
                f"in your sample: {labels}{more}. Is that right?"
            )

        for profile in profiles:
            questions.extend(profile.open_questions)

        enums = [f for f in definition.fields() if f.options]
        if enums:
            preview = "; ".join(f"{f.label} ({', '.join(f.options[:4])})" for f in enums[:3])
            questions.append(
                f"I treated these as picklists: {preview}. "
                "Should any allow values outside those options?"
            )

        missing_rationale = [
            f.label for f in definition.fields() if not f.rationale and f.is_mandatory
        ]
        if missing_rationale:
            questions.append(
                "I couldn't work out why these are needed, so I can't explain them to users "
                f"who ask: {', '.join(missing_rationale[:5])}. What's the reason for each?"
            )

        if len(rows) < 2:
            questions.append(
                "The sample had very little data. Are there fields that only apply in "
                "certain cases, so I can make them conditional?"
            )

        empty_columns = [p.name for p in by_id.values() if p.total and p.filled == 0]
        if empty_columns:
            questions.append(
                f"These columns were entirely blank: {', '.join(empty_columns[:6])}. "
                "Keep them, or drop them from the form?"
            )

        return questions


async def apply_answers(
    definition: FormDefinition,
    instructions: str,
    *,
    llm_provider: Any | None = None,
) -> FormDefinition:
    """Apply a user's free-text corrections to a draft definition.

    This is the second half of the facilitator loop: "make the budget field
    optional and add a field for the rollback plan". Changes are applied
    field-by-field against the existing definition rather than regenerating it,
    so the parts the user was happy with cannot drift.
    """
    from sa_connectors.llm.base import Message

    provider = llm_provider
    if provider is None:
        from sa_connectors.llm import build_provider

        provider = build_provider()

    schema: dict[str, Any] = {
        "type": "object",
        "properties": {
            "field_updates": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "field_id": {"type": "string"},
                        "label": {"type": "string"},
                        "description": {"type": "string"},
                        "rationale": {"type": "string"},
                        "importance": {
                            "type": "string",
                            "enum": ["mandatory", "recommended", "optional"],
                        },
                        "remove": {"type": "boolean"},
                    },
                    "required": ["field_id"],
                    "additionalProperties": False,
                },
            },
            "new_fields": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "field_id": {"type": "string"},
                        "label": {"type": "string"},
                        "type": {"type": "string"},
                        "importance": {
                            "type": "string",
                            "enum": ["mandatory", "recommended", "optional"],
                        },
                        "description": {"type": "string"},
                        "rationale": {"type": "string"},
                        "section_id": {"type": "string"},
                    },
                    "required": ["field_id", "label", "type", "importance"],
                    "additionalProperties": False,
                },
            },
            "form_title": {"type": "string"},
        },
        "required": [],
        "additionalProperties": False,
    }

    current = [
        f'- `{f.id}` "{f.label}" type={f.type.value} importance={f.importance.value}'
        for f in definition.fields()
    ]
    prompt = (
        f"Current form '{definition.title}':\n"
        + "\n".join(current)
        + f"\n\nSections: {', '.join(s.id for s in definition.sections)}"
        + f"\n\nThe user asked for these changes:\n{instructions}\n\n"
        "Return only the changes. Leave everything else untouched."
    )

    changes = await provider.complete_structured(
        [Message.user(prompt)],
        schema,
        system=(
            "You edit form definitions. Apply only what the user asked for. "
            "Never restructure or rename fields they did not mention."
        ),
    )

    updates = {u["field_id"]: u for u in changes.get("field_updates", [])}
    removed = {fid for fid, u in updates.items() if u.get("remove")}

    sections: list[FormSection] = []
    for section in definition.ordered_sections():
        kept: list[FormField] = []
        for field in section.fields:
            if field.id in removed:
                continue
            update = updates.get(field.id)
            if update:
                field = field.model_copy(
                    update={
                        k: v
                        for k, v in {
                            "label": update.get("label"),
                            "description": update.get("description"),
                            "rationale": update.get("rationale"),
                            "importance": (
                                Importance(update["importance"])
                                if update.get("importance")
                                else None
                            ),
                        }.items()
                        if v is not None
                    }
                )
            kept.append(field)
        if kept:
            sections.append(section.model_copy(update={"fields": kept}))

    known_ids = {f.id for s in sections for f in s.fields}
    for new in changes.get("new_fields", []):
        field_id = _slug(new["field_id"])
        if field_id in known_ids:
            continue
        try:
            field = FormField(
                id=field_id,
                label=new["label"],
                type=FieldType(new["type"]),
                importance=Importance(new["importance"]),
                description=new.get("description", ""),
                rationale=new.get("rationale", ""),
            )
        except (ValueError, KeyError) as exc:
            logger.warning(
                "skipping an invalid proposed field",
                extra={"field": new.get("field_id"), "error": str(exc)},
            )
            continue

        target = next(
            (s for s in sections if s.id == new.get("section_id")),
            sections[-1] if sections else None,
        )
        if target is None:
            sections.append(FormSection(id="details", title="Details", fields=[field], order=0))
        else:
            target.fields.append(field)
        known_ids.add(field_id)

    if not sections:
        raise ValidationError("the requested changes would leave the form with no fields")

    return definition.model_copy(
        update={
            "title": changes.get("form_title") or definition.title,
            "sections": sections,
        }
    )


__all__ = [
    "ColumnProfile",
    "FormAuthor",
    "InferenceReport",
    "apply_answers",
    "profile_columns",
    "read_tabular",
]
